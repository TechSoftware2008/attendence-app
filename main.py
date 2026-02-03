import os
import datetime
import holidays
from kivy.app import App
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.button import Button
from kivy.uix.label import Label
from kivy.uix.popup import Popup
from plyer import notification

from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill

from pydrive2.auth import GoogleAuth
from pydrive2.drive import GoogleDrive

FILENAME = "attendance_backup.xlsx"
FOLDER_ID = "10JHVLycrR1A35entOa_nvMHHNJTMsrQM"

# 🇮🇳 Indian Holidays
ind_holidays = holidays.India()

# 🏖 Vacation dates (EDIT IF NEEDED)
VACATIONS = [
    "15-06-2026",
    "16-06-2026"
]

# 🎨 Excel Colors
GREEN = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RED = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
YELLOW = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")


def init_excel():
    if not os.path.exists(FILENAME):
        wb = Workbook()
        ws = wb.active
        ws.title = "Attendance"
        ws.append(["Date", "Day", "Status", "Synced"])
        wb.save(FILENAME)


def already_marked(date_str):
    wb = load_workbook(FILENAME)
    ws = wb.active
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] == date_str:
            return True
    return False


def save_attendance(date_str, day_name, status, synced=False):
    wb = load_workbook(FILENAME)
    ws = wb.active
    ws.append([date_str, day_name, status, str(synced)])

    row = ws.max_row
    if status == "Present":
        ws[f"C{row}"].fill = GREEN
    else:
        ws[f"C{row}"].fill = RED

    if not synced:
        ws[f"D{row}"].fill = YELLOW

    wb.save(FILENAME)


def mark_all_synced():
    wb = load_workbook(FILENAME)
    ws = wb.active
    for row in ws.iter_rows(min_row=2):
        row[3].value = "True"
        row[3].fill = PatternFill(fill_type=None)
    wb.save(FILENAME)


def calculate_attendance():
    wb = load_workbook(FILENAME)
    ws = wb.active
    total = present = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        total += 1
        if row[2] == "Present":
            present += 1

    return (present / total) * 100 if total else 0


def upload_to_drive():
    gauth = GoogleAuth()
    gauth.LocalWebserverAuth()
    drive = GoogleDrive(gauth)

    file_list = drive.ListFile(
        {'q': f"'{FOLDER_ID}' in parents and trashed=false"}
    ).GetList()

    existing = None
    for file in file_list:
        if file['title'] == FILENAME:
            existing = file
            break

    if existing:
        existing.SetContentFile(FILENAME)
        existing.Upload()
    else:
        file_drive = drive.CreateFile({'title': FILENAME, 'parents': [{'id': FOLDER_ID}]})
        file_drive.SetContentFile(FILENAME)
        file_drive.Upload()

    mark_all_synced()


class AttendanceApp(App):
    def build(self):
        init_excel()
        layout = BoxLayout(orientation='vertical', padding=20, spacing=20)

        today = datetime.date.today()
        self.date_str = today.strftime("%d-%m-%Y")
        self.day_name = today.strftime("%A")

        self.info_label = Label(
            text=f"📅 {self.date_str} ({self.day_name})",
            font_size=22
        )

        btn_present = Button(text="Mark Present", background_color=(0, 1, 0, 1))
        btn_absent = Button(text="Mark Absent", background_color=(1, 0, 0, 1))
        btn_sync = Button(text="Sync to Google Drive")
        btn_percent = Button(text="Show Attendance %")

        btn_present.bind(on_press=lambda x: self.mark("Present"))
        btn_absent.bind(on_press=lambda x: self.mark("Absent"))
        btn_sync.bind(on_press=lambda x: self.sync_drive())
        btn_percent.bind(on_press=lambda x: self.show_percent())

        layout.add_widget(Label(text="🎓 Smart Attendance Tracker", font_size=24))
        layout.add_widget(self.info_label)
        layout.add_widget(btn_present)
        layout.add_widget(btn_absent)
        layout.add_widget(btn_sync)
        layout.add_widget(btn_percent)

        return layout

    def mark(self, status):
        today = datetime.date.today()
        date_str = today.strftime("%d-%m-%Y")
        day_name = today.strftime("%A")

        if today.weekday() == 6:
            self.popup("Sunday — No school")
            return

        if today in ind_holidays:
            self.popup("Holiday — No school")
            return

        if date_str in VACATIONS:
            self.popup("Vacation — No attendance needed")
            return

        if already_marked(date_str):
            self.popup("Attendance already marked today")
            return

        save_attendance(date_str, day_name, status, False)
        notification.notify(title="Attendance Saved", message=f"{status} marked")
        self.popup(f"{day_name} marked as {status}")

    def sync_drive(self):
        try:
            upload_to_drive()
            self.popup("Backup Successful to Google Drive")
        except Exception as e:
            self.popup(f"Drive Error: {str(e)}")

    def show_percent(self):
        percent = calculate_attendance()
        self.popup(f"Attendance: {percent:.1f}%")

    def popup(self, msg):
        Popup(title="Info",
              content=Label(text=msg),
              size_hint=(None, None), size=(320, 200)).open()


if __name__ == "__main__":
    AttendanceApp().run()
